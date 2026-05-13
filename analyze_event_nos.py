import openpyxl
import re
from collections import defaultdict

# Read the Excel file
excel_file = r'e:\icddrb projects\ICCRB-Event-DataDashboard\Input files\Stool information of Conjugate Vaccine  (31).xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb['Sheet1']

# Find Event No column
event_no_col_idx = None
for col_idx, cell in enumerate(ws[1], 1):
    if cell.value and 'event' in str(cell.value).lower() and 'no' in str(cell.value).lower():
        event_no_col_idx = col_idx
        print(f"Found Event No Column at index {col_idx}: '{cell.value}'")
        break

if event_no_col_idx:
    # Extract all Event No values
    event_nos = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=event_no_col_idx), start=2):
        cell_value = row[event_no_col_idx - 1].value
        if cell_value and str(cell_value).strip():
            event_nos.append(str(cell_value))
    
    print(f"Total data rows: {ws.max_row - 1}")
    print(f"Total Event No values: {len(event_nos)}")
    print(f"Unique Event No values: {len(set(event_nos))}\n")
    
    # Group by normalized base
    normalize_regex = r',?\s*\(?Day[- ]?\d+\)?\s*'
    
    groups = defaultdict(list)
    for event_no in event_nos:
        normalized = re.sub(normalize_regex, '', event_no, flags=re.IGNORECASE).strip()
        groups[normalized].append(event_no)
    
    # Sort by group size
    sorted_groups = sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)
    
    print("=" * 80)
    print("Event No Groups (by normalized base):\n")
    for base, variations in sorted_groups:
        if base:  # Only show non-empty bases
            unique_variations = sorted(set(variations))
            print(f"Base: '{base}' ({len(variations)} occurrences, {len(unique_variations)} unique)")
            for var in unique_variations[:5]:  # Show first 5 variations
                print(f"  - {var}")
            if len(unique_variations) > 5:
                print(f"  ... and {len(unique_variations) - 5} more variations")
            print()
    
    # Show any empty/problematic cases
    if '' in groups:
        print(f"\n⚠️  EMPTY after normalization ({len(groups[''])} cases):")
        unique_empty = sorted(set(groups['']))[:10]
        for var in unique_empty:
            print(f"  - '{var}'")

wb.close()
