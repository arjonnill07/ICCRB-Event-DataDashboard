import openpyxl
from openpyxl import load_workbook

# File path
file_path = r'e:\icddrb projects\ICCRB-Event-DataDashboard\Input files\Stool information of Conjugate Vaccine  (31).xlsx'

# Load workbook
wb = load_workbook(file_path)

print('=== 1. SHEET NAMES ===')
sheet_names = wb.sheetnames
for i, name in enumerate(sheet_names, 1):
    print(f'{i}. {name}')

# Get first sheet
first_sheet_name = sheet_names[0]
ws = wb[first_sheet_name]

print(f'\n=== 2. COLUMNS IN FIRST SHEET ("{first_sheet_name}") ===')
# Get column headers from first row
headers = []
for col_idx, cell in enumerate(ws[1], 1):
    headers.append(cell.value)
    print(f'{col_idx}. {cell.value}')

print(f'\n=== 3. SEARCHING FOR PARTICIPANT R06107 ===')
# Search all cells for R06107
found = False
matching_data = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    for cell_val in row:
        if cell_val and isinstance(cell_val, str) and 'R06107' in cell_val.upper():
            matching_data.append(row)
            found = True
            break

if found:
    print(f'Found {len(matching_data)} matching rows:')
    for idx, row in enumerate(matching_data):
        print(f'\nRow {idx + 1}:')
        for col_idx, (header, value) in enumerate(zip(headers, row), 1):
            print(f'  {header}: {value}')
else:
    print('Participant R06107 not found')

print(f'\n=== 4. SAMPLE ROWS (first 3 rows) ===')
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), 1):
    print(f'Row {row_idx}: {row}')

print(f'\n=== SUMMARY ===')
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}')
